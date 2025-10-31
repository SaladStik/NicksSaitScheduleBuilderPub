import streamlit as st
import json
import time
from datetime import datetime, timedelta
from itertools import combinations
import random
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from icalendar import Calendar, Event
import pytz
import re
import requests
from typing import List, Dict
from parse_headers import parse_request_headers, format_for_banner_api

# Initialize session state for browser-based storage (stateless)
if "classes_data" not in st.session_state:
    st.session_state.classes_data = {}

# BANNER API FUNCTIONS


def get_banner_credentials():
    """Get Banner API credentials from Streamlit secrets or session state"""
    if "banner_cookies" not in st.session_state:
        # Try to get from secrets first
        try:
            st.session_state.banner_cookies = {
                "JSESSIONID": st.secrets["BANNER"]["JSESSIONID"],
                "NLB": st.secrets["BANNER"]["NLB"],
                "NSC_ESNS": st.secrets["BANNER"]["NSC_ESNS"],
            }
            st.session_state.banner_token = st.secrets["BANNER"]["SYNC_TOKEN"]
        except:
            # No credentials available - user needs to authenticate
            return None, None

    return st.session_state.banner_cookies, st.session_state.banner_token


def fetch_available_terms() -> List[Dict]:
    """
    Fetch available terms from Banner API

    Returns:
        List of term dictionaries with 'code' and 'description'
    """
    import time

    cookies, sync_token = get_banner_credentials()

    if not cookies or not sync_token:
        st.error("❌ No authentication credentials found.")
        return []

    terms_url = "https://sait-sust-prd-prd1-ban-ss-ssag6.sait.ca/StudentRegistrationSsb/ssb/classRegistration/getTerms"

    params = {"searchTerm": "", "offset": 1, "max": 10, "_": int(time.time() * 1000)}

    headers = {
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Accept-Encoding": "gzip, deflate, br, zstd",
        "Accept-Language": "en-US,en;q=0.9",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
        "Referer": "https://sait-sust-prd-prd1-ban-ss-ssag6.sait.ca/StudentRegistrationSsb/ssb/term/termSelection?mode=registration",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin",
        "X-Requested-With": "XMLHttpRequest",
        "X-Synchronizer-Token": sync_token,
        "sec-ch-ua": '"Chromium";v="140", "Not=A?Brand";v="24", "Google Chrome";v="140"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
    }

    try:
        response = requests.get(
            terms_url, params=params, headers=headers, cookies=cookies, timeout=15
        )

        if response.status_code == 200:
            terms = response.json()
            return terms if isinstance(terms, list) else []
        else:
            st.error(f"Failed to fetch terms: HTTP {response.status_code}")
            return []

    except Exception as e:
        st.error(f"Error fetching terms: {e}")
        return []


def search_courses(search_term: str, term: str = "202530") -> List[Dict]:
    """
    Search for courses using the Banner API autocomplete

    Args:
        search_term: The search term (e.g., "abdy", "itsc", "cprg")
        term: Term code (e.g., "202530")

    Returns:
        List of course dictionaries with 'code' and 'description'
    """
    import time

    cookies, sync_token = get_banner_credentials()

    # Check if credentials are available
    if not cookies or not sync_token:
        st.error("❌ No authentication credentials found. Please authenticate first.")
        return []

    search_url = "https://sait-sust-prd-prd1-ban-ss-ssag6.sait.ca/StudentRegistrationSsb/ssb/classSearch/get_subjectcoursecombo"

    search_params = {
        "searchTerm": search_term,
        "term": term,
        "offset": 1,
        "max": 500,  # Get more results
    }

    headers = {
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Accept-Encoding": "gzip, deflate, br, zstd",
        "Accept-Language": "en-US,en;q=0.9",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
        "Referer": "https://sait-sust-prd-prd1-ban-ss-ssag6.sait.ca/StudentRegistrationSsb/ssb/classRegistration/classRegistration",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36",
        "X-Requested-With": "XMLHttpRequest",
        "X-Synchronizer-Token": sync_token,
        "sec-ch-ua": '"Chromium";v="140", "Not=A?Brand";v="24", "Google Chrome";v="140"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
    }

    try:
        response = requests.get(
            search_url,
            params=search_params,
            headers=headers,
            cookies=cookies,
            timeout=15,
        )

        if response.status_code == 200:
            data = response.json()
            # Return list of course objects
            return data if isinstance(data, list) else []
        else:
            st.error(f"Search failed: HTTP {response.status_code}")
            return []

    except Exception as e:
        st.error(f"Error searching courses: {e}")
        return []


def reset_banner_search(term: str = "202530") -> bool:
    """Reset the Banner search state"""
    cookies, sync_token = get_banner_credentials()

    # Check if credentials are available
    if not cookies or not sync_token:
        return False

    reset_url = "https://sait-sust-prd-prd1-ban-ss-ssag6.sait.ca/StudentRegistrationSsb/ssb/classSearch/resetDataForm"

    headers = {
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
        "Referer": "https://sait-sust-prd-prd1-ban-ss-ssag6.sait.ca/StudentRegistrationSsb/ssb/classRegistration/classRegistration",
        "X-Synchronizer-Token": sync_token,
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    }

    try:
        response = requests.post(
            reset_url, headers=headers, cookies=cookies, data={"term": term}, timeout=10
        )
        return response.status_code == 200
    except:
        return False


def fetch_banner_api(
    term: str, course_code: str, session_id: str = None, open_only: bool = True
) -> Dict:
    """
    Fetch course data from SAIT Banner API (with authentication)

    Args:
        term: Term code (e.g., "202530" for Winter 2026)
        course_code: Course code (e.g., "ITSC320", "CPSY300")
        session_id: Optional unique session ID (will generate if not provided)
        open_only: If True, only fetch classes with available seats

    Returns:
        Dictionary with API response
    """
    import time

    cookies, sync_token = get_banner_credentials()

    # Check if credentials are available
    if not cookies or not sync_token:
        st.error("❌ No authentication credentials found. Please authenticate first.")
        return {"success": False, "error": "No credentials"}

    # Generate unique session ID if not provided
    if not session_id:
        session_id = f"streamlit{int(time.time() * 1000)}"

    # Step 1: Search endpoint to set up what we're searching for
    search_url = "https://sait-sust-prd-prd1-ban-ss-ssag6.sait.ca/StudentRegistrationSsb/ssb/classSearch/get_subjectcoursecombo"

    search_params = {"searchTerm": course_code, "term": term, "offset": 1, "max": 10}

    headers = {
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Accept-Encoding": "gzip, deflate, br, zstd",
        "Accept-Language": "en-US,en;q=0.9",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
        "Referer": "https://sait-sust-prd-prd1-ban-ss-ssag6.sait.ca/StudentRegistrationSsb/ssb/classRegistration/classRegistration",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36",
        "X-Requested-With": "XMLHttpRequest",
        "X-Synchronizer-Token": sync_token,
        "sec-ch-ua": '"Chromium";v="140", "Not=A?Brand";v="24", "Google Chrome";v="140"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
    }

    try:
        # Step 1: Search
        search_response = requests.get(
            search_url,
            params=search_params,
            headers=headers,
            cookies=cookies,
            timeout=15,
        )

        if search_response.status_code != 200:
            st.error(
                f"Search failed for {course_code}: HTTP {search_response.status_code}"
            )
            return None

        search_data = search_response.json()
        if not search_data or len(search_data) == 0:
            st.warning(f"No course found matching {course_code}")
            return None

        # Get the actual course code from search results
        actual_code = search_data[0]["code"]

        # Step 2: Get results
        results_url = "https://sait-sust-prd-prd1-ban-ss-ssag6.sait.ca/StudentRegistrationSsb/ssb/searchResults/searchResults"

        results_params = {
            "txt_subjectcoursecombo": actual_code,
            "txt_term": term,
            "startDatepicker": "",
            "endDatepicker": "",
            "uniqueSessionId": session_id,
            "pageOffset": 0,
            "pageMaxSize": 50,
            "sortColumn": "subjectDescription",
            "sortDirection": "asc",
        }

        # Only add chk_open_only if filtering by available seats
        if open_only:
            results_params["chk_open_only"] = "true"

        results_response = requests.get(
            results_url,
            params=results_params,
            headers=headers,
            cookies=cookies,
            timeout=15,
        )
        results_response.raise_for_status()
        data = results_response.json()

        return data

    except requests.exceptions.RequestException as e:
        st.error(f"Error fetching {course_code} from Banner API: {e}")
        return None
    except json.JSONDecodeError as e:
        st.error(f"Failed to parse JSON response for {course_code}: {e}")
        return None


def parse_banner_response(api_response: Dict, open_only: bool = True) -> List[Dict]:
    """
    Parse Banner API response and convert to app's class format

    Args:
        api_response: API response dictionary
        open_only: If True, only include classes with available seats

    Returns:
        List of class dictionaries in app format
    """
    if not api_response:
        st.warning("No response received from API")
        return []

    if not isinstance(api_response, dict):
        st.error(f"Invalid API response format: {type(api_response)}")
        return []

    if not api_response.get("success"):
        st.error(f"API returned error: {api_response.get('message', 'Unknown error')}")
        return []

    classes = []
    data_items = api_response.get("data", [])

    # Additional safety check - ensure data_items is not None
    if data_items is None:
        st.warning("No data returned from API")
        return []

    for course_data in data_items:
        seats_available = course_data.get("seatsAvailable", 0)

        # Only include courses with available seats if open_only is True
        if open_only and seats_available <= 0:
            continue

        # Parse meeting times
        schedule = []
        for meeting in course_data.get("meetingsFaculty", []):
            meeting_time = meeting.get("meetingTime", {})

            # Skip if no meeting time data
            if not meeting_time:
                continue

            # Get times
            begin_time = meeting_time.get("beginTime")
            end_time = meeting_time.get("endTime")

            if not begin_time or not end_time:
                continue

            # Convert military time (e.g., "0800") to HH:MM format
            start_time = f"{begin_time[:2]}:{begin_time[2:]}"
            end_time_str = f"{end_time[:2]}:{end_time[2:]}"

            # Get location
            building = meeting_time.get("building", "")
            room = meeting_time.get("room", "")
            building_room = f"{building}{room}" if building or room else "TBA"

            # Map day booleans to day names
            day_mapping = {
                "monday": "Monday",
                "tuesday": "Tuesday",
                "wednesday": "Wednesday",
                "thursday": "Thursday",
                "friday": "Friday",
                "saturday": "Saturday",
                "sunday": "Sunday",
            }

            # Add schedule entry for each day that is True
            for day_key, day_name in day_mapping.items():
                if meeting_time.get(day_key, False):
                    schedule.append(
                        {
                            "day": day_name,
                            "start_time": start_time,
                            "end_time": end_time_str,
                            "class_room": building_room,
                        }
                    )

        if schedule:  # Only add if there's a valid schedule
            # Get instructor name
            instructor = "TBA"
            if course_data.get("faculty"):
                instructor = course_data["faculty"][0].get("displayName", "TBA")

            class_obj = {
                "name": f"{course_data['subject']} {course_data['courseNumber']}",
                "group": course_data.get("sequenceNumber", "A"),
                "schedule": schedule,
                "seats_available": course_data.get("seatsAvailable", 0),
                "max_enrollment": course_data.get("maximumEnrollment", 0),
                "instructor": instructor,
                "crn": course_data.get("courseReferenceNumber", ""),
            }
            classes.append(class_obj)
        else:
            # No valid schedule found - skip this class
            pass

    return classes


def fetch_all_available_courses(
    term: str, course_codes: List[str], open_only: bool = True
) -> List[Dict]:
    """
    Fetch all available courses for given course codes

    Args:
        term: Term code (e.g., "202530")
        course_codes: List of course codes (e.g., ["ITSC320", "CPSY300", "INTP302"])
        open_only: If True, only fetch classes with available seats

    Returns:
        List of all available classes
    """
    import time

    all_classes = []
    session_id = f"streamlit{int(time.time() * 1000)}"

    progress_bar = st.progress(0)
    status_text = st.empty()

    for idx, course_code in enumerate(course_codes):
        # Reset search state before each course
        reset_banner_search(term)
        time.sleep(0.5)

        status_text.text(
            f"Fetching {course_code} courses... ({idx+1}/{len(course_codes)})"
        )

        api_response = fetch_banner_api(term, course_code, session_id, open_only)
        if api_response:
            classes = parse_banner_response(api_response, open_only=open_only)
            all_classes.extend(classes)
            st.success(
                f"✅ Found {len(classes)} available section(s) for {course_code}"
            )
        else:
            st.warning(f"⚠️ No data returned for {course_code}")

        # Update progress
        progress_bar.progress((idx + 1) / len(course_codes))

        # Small delay to avoid overwhelming the server
        time.sleep(1)

    progress_bar.empty()
    status_text.empty()

    return all_classes


# REGISTRATION API FUNCTIONS


def get_current_registrations(term: str) -> List[Dict]:
    """
    Get user's current registered classes for a term

    Args:
        term: Term code (e.g., "202530")

    Returns:
        List of registered course objects
    """
    cookies, sync_token = get_banner_credentials()

    if not cookies or not sync_token:
        st.error("❌ No authentication credentials found.")
        return []

    # Note: termFilter should be EMPTY to get all registrations, then we filter by term
    url = f"https://sait-sust-prd-prd1-ban-ss-ssag6.sait.ca/StudentRegistrationSsb/ssb/classRegistration/getRegistrationEvents?termFilter="

    headers = {
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Accept-Encoding": "gzip, deflate, br, zstd",
        "Accept-Language": "en-US,en;q=0.9",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
        "Referer": "https://sait-sust-prd-prd1-ban-ss-ssag6.sait.ca/StudentRegistrationSsb/ssb/classRegistration/classRegistration",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin",
        "X-Requested-With": "XMLHttpRequest",
        "X-Synchronizer-Token": sync_token,
        "sec-ch-ua": '"Chromium";v="140", "Not=A?Brand";v="24", "Google Chrome";v="140"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
    }

    try:
        response = requests.get(url, headers=headers, cookies=cookies, timeout=15)

        if response.status_code == 200:
            try:
                data = response.json()

                if isinstance(data, list):
                    # Filter by term if specified
                    if term:
                        filtered_data = [
                            item for item in data if item.get("term") == term
                        ]
                        data = filtered_data

                return data if isinstance(data, list) else []
            except Exception as e:
                # Response is not JSON (might be HTML error page)
                st.warning("⚠️ Could not parse registration data")
                return []
        else:
            # Don't show error for 500 - just return empty list
            if response.status_code != 500:
                st.error(f"Failed to get registrations: HTTP {response.status_code}")
            else:
                st.error(
                    "⚠️ HTTP 500 - Session may be expired. Try updating your authentication tokens!"
                )
            return []

    except Exception as e:
        # Silently handle errors - user might not have any registrations
        return []


def get_registration_models_for_term(term: str) -> Dict[str, Dict]:
    """
    Extract full registration model objects from the Banner registration page HTML.
    The models are embedded in window.bootstraps.summaryModels in the JavaScript.
    These models contain all the required fields that Banner expects for drop operations.

    Args:
        term: Term code

    Returns:
        Dictionary of {crn: model_object}
    """
    cookies, sync_token = get_banner_credentials()

    if not cookies or not sync_token:
        return {}

    # Fetch the registration page HTML which contains the models
    url = "https://sait-sust-prd-prd1-ban-ss-ssag6.sait.ca/StudentRegistrationSsb/ssb/classRegistration/classRegistration"

    headers = {
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
        "Accept-Encoding": "gzip, deflate, br, zstd",
        "Accept-Language": "en-US,en;q=0.9",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
        "Referer": "https://sait-sust-prd-prd1-ban-ss-ssag6.sait.ca/StudentRegistrationSsb/ssb/term/termSelection?mode=registration",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "same-origin",
        "Upgrade-Insecure-Requests": "1",
        "sec-ch-ua": '"Chromium";v="140", "Not=A?Brand";v="24", "Google Chrome";v="140"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
    }

    try:
        response = requests.get(url, headers=headers, cookies=cookies, timeout=15)

        if response.status_code == 200:
            html = response.text

            # Extract the summaryModels from the JavaScript in the HTML
            # Look for: summaryModels: [...] in window.bootstraps
            import re
            import json

            # Find the start of summaryModels array
            start_pattern = r"summaryModels:\s*\["
            match = re.search(start_pattern, html)

            if match:
                start_pos = match.end() - 1  # Position of the opening '['

                # Now manually find the matching closing bracket by counting
                bracket_count = 0
                end_pos = start_pos
                in_string = False
                escape_next = False

                for i in range(start_pos, len(html)):
                    char = html[i]

                    # Handle string escaping
                    if escape_next:
                        escape_next = False
                        continue

                    if char == "\\":
                        escape_next = True
                        continue

                    # Track if we're inside a string
                    if char == '"' and not escape_next:
                        in_string = not in_string
                        continue

                    # Only count brackets outside of strings
                    if not in_string:
                        if char == "[":
                            bracket_count += 1
                        elif char == "]":
                            bracket_count -= 1
                            if bracket_count == 0:
                                end_pos = i + 1
                                break

                if bracket_count == 0:
                    models_json = html[start_pos:end_pos]

                    # Parse the JSON array
                    try:
                        models = json.loads(models_json)

                        # Build dictionary of CRN -> model
                        models_by_crn = {}
                        for model in models:
                            if (
                                isinstance(model, dict)
                                and "courseReferenceNumber" in model
                            ):
                                crn = str(model["courseReferenceNumber"])
                                models_by_crn[crn] = model

                        return models_by_crn
                    except json.JSONDecodeError as e:
                        return {}

            return {}

        return {}

    except Exception as e:
        return {}


def add_class_to_cart(term: str, crn: str) -> Dict:
    """
    Add a class to registration cart

    Args:
        term: Term code (e.g., "202530")
        crn: Course Reference Number

    Returns:
        Dict with 'success' boolean and 'data' or 'error' message
    """
    cookies, sync_token = get_banner_credentials()

    if not cookies or not sync_token:
        return {"success": False, "error": "No authentication credentials"}

    url = f"https://sait-sust-prd-prd1-ban-ss-ssag6.sait.ca/StudentRegistrationSsb/ssb/classRegistration/addRegistrationItem?term={term}&courseReferenceNumber={crn}&olr=false"

    headers = {
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Accept-Encoding": "gzip, deflate, br, zstd",
        "Accept-Language": "en-US,en;q=0.9",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
        "Referer": "https://sait-sust-prd-prd1-ban-ss-ssag6.sait.ca/StudentRegistrationSsb/ssb/classRegistration/classRegistration",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin",
        "X-Requested-With": "XMLHttpRequest",
        "X-Synchronizer-Token": sync_token,
        "sec-ch-ua": '"Chromium";v="140", "Not=A?Brand";v="24", "Google Chrome";v="140"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
    }

    try:
        response = requests.get(url, headers=headers, cookies=cookies, timeout=15)

        if response.status_code == 200:
            data = response.json()
            if data.get("success", False):
                return {"success": True, "data": data}
            else:
                # Get error message from response
                error_msg = data.get("message", "Unknown error")
                return {"success": False, "error": error_msg}
        else:
            return {"success": False, "error": f"HTTP {response.status_code}"}

    except Exception as e:
        return {"success": False, "error": str(e)}


def submit_registration(term: str, registration_items: List[Dict]) -> Dict:
    """
    Submit registration changes (add/drop classes)

    Args:
        term: Term code
        registration_items: List of registration item objects to update

    Returns:
        Response dictionary
    """
    cookies, sync_token = get_banner_credentials()

    if not cookies or not sync_token:
        return {"success": False, "error": "No credentials"}

    url = "https://sait-sust-prd-prd1-ban-ss-ssag6.sait.ca/StudentRegistrationSsb/ssb/classRegistration/submitRegistration/batch"

    headers = {
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Accept-Encoding": "gzip, deflate, br, zstd",
        "Accept-Language": "en-US,en;q=0.9",
        "Cache-Control": "no-cache",
        "Origin": "https://sait-sust-prd-prd1-ban-ss-ssag6.sait.ca",
        "Pragma": "no-cache",
        "Referer": "https://sait-sust-prd-prd1-ban-ss-ssag6.sait.ca/StudentRegistrationSsb/ssb/classRegistration/classRegistration",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin",
        "X-Requested-With": "XMLHttpRequest",
        "X-Synchronizer-Token": sync_token,
        "sec-ch-ua": '"Chromium";v="140", "Not=A?Brand";v="24", "Google Chrome";v="140"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
    }

    unique_session_id = f"streamlit{int(time.time() * 1000)}"

    payload = {
        "create": [],
        "update": registration_items,
        "destroy": [],
        "uniqueSessionId": unique_session_id,
    }

    try:
        response = requests.post(
            url, headers=headers, cookies=cookies, json=payload, timeout=15
        )

        if response.status_code == 200:
            return response.json()
        else:
            return {"success": False, "error": f"HTTP {response.status_code}"}

    except Exception as e:
        return {"success": False, "error": str(e)}


def drop_classes(term: str, crns: List[str]) -> Dict:
    """
    Drop classes from registration
    Uses the registration models stored in session state from manual_schedule_editor

    Args:
        term: Term code
        crns: List of CRNs to drop

    Returns:
        Response dictionary with success status
    """
    cookies, sync_token = get_banner_credentials()

    if not cookies or not sync_token:
        return {"success": False, "error": "No credentials"}

    # Get registration models from session state
    registration_models = st.session_state.get("registration_models", {})

    if not registration_models:
        st.info(
            "� This shouldn't happen - models should be loaded when schedule is displayed"
        )
        return {
            "success": False,
            "error": "No registration models available. Please refresh the page.",
        }

    # Find the models for the CRNs we want to drop
    models_to_drop = []
    for crn in crns:
        if crn in registration_models:
            model = registration_models[
                crn
            ].copy()  # Make a copy to avoid modifying session state
            # Set the action to DROP
            model["selectedAction"] = "DW"
            models_to_drop.append(model)

    if not models_to_drop:
        return {
            "success": False,
            "error": "Could not find registration models for specified CRNs",
        }

    # Step 2: Submit the drop via batch endpoint
    url = "https://sait-sust-prd-prd1-ban-ss-ssag6.sait.ca/StudentRegistrationSsb/ssb/classRegistration/submitRegistration/batch"

    headers = {
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Accept-Encoding": "gzip, deflate, br, zstd",
        "Accept-Language": "en-US,en;q=0.9",
        "Cache-Control": "no-cache",
        "Content-Type": "application/json",
        "Origin": "https://sait-sust-prd-prd1-ban-ss-ssag6.sait.ca",
        "Pragma": "no-cache",
        "Referer": "https://sait-sust-prd-prd1-ban-ss-ssag6.sait.ca/StudentRegistrationSsb/ssb/classRegistration/classRegistration",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin",
        "X-Requested-With": "XMLHttpRequest",
        "X-Synchronizer-Token": sync_token,
        "sec-ch-ua": '"Chromium";v="140", "Not=A?Brand";v="24", "Google Chrome";v="140"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
    }

    unique_session_id = f"streamlit{int(time.time() * 1000)}"

    # Build payload with update array containing models with selectedAction='DW'
    payload = {
        "create": [],
        "update": models_to_drop,  # Send the full models with selectedAction='DW'
        "destroy": [],  # Keep empty as per real API call
        "uniqueSessionId": unique_session_id,
    }

    st.info(f"  - create: {len(payload['create'])} items")
    st.info(f"  - update: {len(payload['update'])} items")
    st.info(f"  - destroy: {len(payload['destroy'])} items")
    st.info(f"  - uniqueSessionId: {unique_session_id}")

    try:
        response = requests.post(
            url, headers=headers, cookies=cookies, json=payload, timeout=15
        )

        if response.status_code == 200:
            try:
                data = response.json()
                return {"success": True, "data": data}
            except Exception as e:
                return {"success": False, "error": f"JSON parse error: {str(e)}"}
        else:
            return {"success": False, "error": f"HTTP {response.status_code}"}

    except Exception as e:
        return {"success": False, "error": str(e)}


def apply_schedule_to_banner(schedule_combo: List[Dict], term: str) -> bool:
    """
    Apply a schedule combination to Banner (drop existing classes, then register new ones one by one)

    Args:
        schedule_combo: List of classes in the selected schedule
        term: Term code

    Returns:
        True if successful, False otherwise
    """
    # Step 0: Get current registrations and drop them all
    st.info("🗑️ Step 1: Dropping all current classes...")

    registration_models = st.session_state.get("registration_models", {})
    current_crns = list(registration_models.keys())

    failed_drops = []
    dropped_count = 0

    if current_crns:
        st.info(f"Found {len(current_crns)} currently registered class(es)")

        # Try to drop each class individually
        for crn in current_crns:
            model = registration_models.get(crn, {})
            course_name = f"{model.get('subject', '')} {model.get('courseNumber', '')} (CRN: {crn})"

            st.info(f"Dropping {course_name}...")
            result = drop_classes(term, [crn])

            if result.get("success", False):
                st.success(f"✅ Dropped {course_name}")
                dropped_count += 1
            else:
                error_msg = result.get("error", "Unknown error")
                st.warning(f"⚠️ Failed to drop {course_name}: {error_msg}")
                failed_drops.append(crn)

            time.sleep(0.5)  # Small delay between drops

        if dropped_count > 0:
            st.success(f"✅ Successfully dropped {dropped_count} class(es)")
        if failed_drops:
            st.warning(
                f"⚠️ {len(failed_drops)} class(es) could not be dropped - will retry after registration"
            )
    else:
        st.info("No currently registered classes to drop")

    # Step 1: Register new classes one by one
    st.info("➕ Step 2: Registering new classes one by one...")
    registered_count = 0
    failed_classes = []

    for cls in schedule_combo:
        crn = cls.get("crn", "")
        if not crn:
            st.warning(
                f"⚠️ No CRN found for {cls.get('name', 'Unknown')} - {cls.get('group', '')}"
            )
            failed_classes.append(
                f"{cls.get('name', 'Unknown')} ({cls.get('group', '')})"
            )
            continue

        course_name = f"{cls.get('name', 'Unknown')} - Section {cls.get('group', '')} (CRN: {crn})"
        st.info(f"Registering {course_name}...")

        # Add to cart
        add_result = add_class_to_cart(term, crn)
        if not add_result.get("success", False):
            error_msg = add_result.get("error", "Unknown error")
            st.error(f"❌ Failed to add {course_name}: {error_msg}")
            failed_classes.append(
                f"{cls.get('name', 'Unknown')} ({cls.get('group', '')})"
            )
            time.sleep(0.5)
            continue

        # Submit registration for this single class
        added_item = add_result.get("data")
        if added_item and "model" in added_item:
            model = added_item["model"]
            model["selectedAction"] = "RB"  # RB = Re-Register (submit)
            model["recordStatus"] = "Q"  # Q = Queued for submission

            submit_result = submit_registration(term, [model])

            if submit_result.get("success", False):
                st.success(f"✅ Registered for {course_name}")
                registered_count += 1
            else:
                error_msg = submit_result.get("error", "Unknown error")
                st.error(f"❌ Failed to register for {course_name}: {error_msg}")
                failed_classes.append(
                    f"{cls.get('name', 'Unknown')} ({cls.get('group', '')})"
                )
        else:
            st.error(f"❌ Failed to prepare {course_name} for registration")
            failed_classes.append(
                f"{cls.get('name', 'Unknown')} ({cls.get('group', '')})"
            )

        time.sleep(0.5)  # Small delay between registrations

    # Step 2: Retry failed drops (if any)
    if failed_drops:
        st.info("🔄 Step 3: Retrying failed drops...")
        retry_dropped = 0
        still_failed = []

        for crn in failed_drops:
            model = registration_models.get(crn, {})
            course_name = f"{model.get('subject', '')} {model.get('courseNumber', '')} (CRN: {crn})"

            st.info(f"Retrying drop for {course_name}...")
            result = drop_classes(term, [crn])

            if result.get("success", False):
                st.success(f"✅ Successfully dropped {course_name}")
                retry_dropped += 1
            else:
                error_msg = result.get("error", "Unknown error")
                st.error(f"❌ Still cannot drop {course_name}: {error_msg}")
                still_failed.append(crn)

            time.sleep(0.5)

        if retry_dropped > 0:
            st.success(f"✅ Successfully dropped {retry_dropped} class(es) on retry")
        if still_failed:
            st.error(f"❌ {len(still_failed)} class(es) still could not be dropped")

    # Summary
    st.markdown("---")
    st.markdown("### 📊 Registration Summary")

    if registered_count > 0:
        st.success(f"✅ Successfully registered for {registered_count} class(es)")

    if failed_classes:
        st.error(f"❌ Failed to register for {len(failed_classes)} class(es):")
        for fail in failed_classes:
            st.error(f"  • {fail}")

    if registered_count > 0:
        st.balloons()
        st.info("🔍 Please check your SAIT Banner account to verify your registration")
        return True
    else:
        st.error("❌ No classes were successfully registered")
        return False


# CLASS LOGGER FUNCTIONS


# Function to display a single class entry
def display_class_entry(class_entry):
    st.write(f"Class name: {class_entry['name'].capitalize()}")
    st.write(f"Group/Section: {class_entry['group']}")
    for schedule in class_entry["schedule"]:
        st.write(
            f"Day: {schedule['day']}, Start time: {schedule['start_time']}, End time: {schedule['end_time']}"
        )


def save_class_to_db(id, class_data):
    """Save class data to browser session state (stateless)"""
    if id not in st.session_state.classes_data:
        st.session_state.classes_data[id] = []
    st.session_state.classes_data[id].append(class_data)


def get_classes_from_db(id):
    """Get classes from browser session state (stateless)"""
    return st.session_state.classes_data.get(id, [])


# ----------------------------

# TIMETABLE CREATOR FUNCTIONS


# Additional functions for timetable
def parse_time(time_str):
    return datetime.strptime(time_str, "%H:%M").time()


def times_overlap(start1, end1, start2, end2):
    return max(start1, start2) < min(end1, end2)


def has_conflict(class1, class2):
    for session1 in class1["schedule"]:
        for session2 in class2["schedule"]:
            if session1["day"] == session2["day"]:
                if times_overlap(
                    session1["start_time"],
                    session1["end_time"],
                    session2["start_time"],
                    session2["end_time"],
                ):
                    return True
    return False


def has_unique_classes(combination):
    class_names = [cls["name"] for cls in combination]
    return len(class_names) == len(set(class_names))


def has_free_days(combination, free_days):
    # Checks if the combination respects the free days
    scheduled_days = {
        session["day"] for cls in combination for session in cls["schedule"]
    }
    return all(day not in scheduled_days for day in free_days)


def get_unique_time_slots(classes):
    time_slots = set()
    for cls in classes:
        for session in cls["schedule"]:
            time_slots.add((session["start_time"], session["end_time"]))
    return sorted(time_slots, key=lambda x: x[0])


def get_random_light_color():
    return "{:02x}{:02x}{:02x}".format(
        random.randint(100, 255), random.randint(100, 255), random.randint(100, 255)
    )


def display_timetable_html(combo, time_slots, classes):
    """Generate HTML for a single timetable"""
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

    # Generate random colors for each class
    class_colors = {cls["name"]: f"#{get_random_light_color()}" for cls in classes}

    # Start building the HTML table
    html = """
    <style>
        .schedule-table {
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
            font-size: 14px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        .schedule-table th {
            background-color: #FFD700;
            color: #333;
            padding: 12px;
            text-align: center;
            border: 1px solid #404040;
            font-weight: bold;
        }
        .schedule-table td {
            padding: 10px;
            border: 1px solid #404040;
            text-align: center;
            vertical-align: middle;
        }
        .time-cell {
            background-color: #FFD700;
            font-weight: bold;
            white-space: nowrap;
        }
        .class-cell {
            font-size: 12px;
            line-height: 1.4;
        }
    </style>
    <table class="schedule-table">
        <tr>
            <th>Time</th>
    """

    for day in days:
        html += f"<th>{day}</th>"
    html += "</tr>"

    # Add rows for each time slot
    for time_slot in time_slots:
        start_str = time_slot[0].strftime("%H:%M")
        end_str = time_slot[1].strftime("%H:%M")
        html += f'<tr><td class="time-cell">{start_str} - {end_str}</td>'

        for day in days:
            cell_content = ""
            for cls in combo:
                for session in cls["schedule"]:
                    if (
                        session["day"] == day
                        and (session["start_time"], session["end_time"]) == time_slot
                    ):
                        bg_color = class_colors[cls["name"]]
                        class_room = session.get("class_room", "N/A")
                        cell_content = f'<div style="background-color: {bg_color}; padding: 8px; border-radius: 4px;" class="class-cell"><strong>{cls["name"]}</strong><br>Group {cls["group"]}<br>Room: {class_room}</div>'
                        break

            html += f"<td>{cell_content}</td>"

        html += "</tr>"

    html += "</table>"
    return html


def create_single_sheet_xlsx_timetables(combinations, filename, time_slots, classes):
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedules"

    header_color = PatternFill(
        start_color="FFD700", end_color="FFD700", fill_type="solid"
    )
    class_colors = {
        cls["name"]: PatternFill(
            start_color=get_random_light_color(),
            end_color=get_random_light_color(),
            fill_type="solid",
        )
        for cls in classes
    }

    current_row = 1
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

    for i, combo in enumerate(combinations):
        current_row += 1  # Move to the next row for the header

        # Set the header row and column
        for col, day in enumerate(["Time"] + days, start=1):
            cell = ws.cell(row=current_row, column=col, value=day)
            if col != 1:  # Apply header color to days, not to the 'Time' column
                cell.fill = header_color

        # Create a dark grey border style
        dark_grey_side = Side(border_style="thin", color="404040")
        dark_grey_border = Border(
            left=dark_grey_side,
            right=dark_grey_side,
            top=dark_grey_side,
            bottom=dark_grey_side,
        )

        # Apply the border to the header row
        for col in range(1, 7):  # Columns A-G
            ws.cell(row=current_row, column=col).border = dark_grey_border

        # Populate the timetable
        for time_slot in time_slots:
            current_row += 1
            ws.cell(
                row=current_row,
                column=1,
                value=f"{time_slot[0].strftime('%H:%M')} - {time_slot[1].strftime('%H:%M')}",
            )
            ws.cell(row=current_row, column=1).fill = header_color

            for cls in combo:
                for session in cls["schedule"]:
                    if (session["start_time"], session["end_time"]) == time_slot:
                        day_col = days.index(session["day"]) + 2
                        cell = ws.cell(row=current_row, column=day_col)
                        class_room_info = session.get("class_room", "N/A")
                        cell.value = f"{cls['name']}\n(Group {cls['group']}\nRoom: {class_room_info})"
                        cell.fill = class_colors[cls["name"]]
                        cell.alignment = Alignment(wrap_text=True)

        # Apply the border to each cell in the timetable
        for row in ws.iter_rows(
            min_row=current_row - len(time_slots),
            max_row=current_row,
            min_col=1,
            max_col=6,
        ):
            for cell in row:
                cell.border = dark_grey_border

        current_row += 5

    # Set column widths
    for i, column_width in enumerate([20] + [25] * 5, start=1):
        ws.column_dimensions[get_column_letter(i)].width = column_width

    wb.save(filename)


# ----------------------------

# CALENDAR ICS GENERATOR FUNCTIONS


def generate_ics_file_for_classes(
    selected_classes,
    classes,
    start_date_str,
    end_date_str,
    filename="class_schedule.ics",
):
    cal = Calendar()

    start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
    end_date = datetime.strptime(end_date_str, "%Y-%m-%d")

    # Day name mapping (UI uses English day names now)
    day_mapping = {
        "Monday": "Monday",
        "Tuesday": "Tuesday",
        "Wednesday": "Wednesday",
        "Thursday": "Thursday",
        "Friday": "Friday",
        "Saturday": "Saturday",
        "Sunday": "Sunday",
    }

    for selected_class in selected_classes:

        cls = next(
            (
                item
                for item in classes
                if item["name"] == selected_class["name"]
                and item["group"] == selected_class["group"]
            ),
            None,
        )

        if cls is None:
            continue

        # Get instructor if available
        instructor = cls.get("instructor", "TBA")

        for session in cls["schedule"]:
            day, start_time_str, end_time_str = (
                session["day"],
                session["start_time"],
                session["end_time"],
            )
            if not all([day, start_time_str, end_time_str]):
                continue

            # Translate day name to English
            day_english = day_mapping.get(day, day)

            first_occurrence_date = start_date
            # Adjust first_occurrence_date to the first occurrence of the session day
            while first_occurrence_date.strftime("%A") != day_english:
                first_occurrence_date += timedelta(days=1)
                if first_occurrence_date > end_date:
                    # If first_occurrence_date exceeds end_date, break the loop
                    break

            if first_occurrence_date > end_date:
                # If the first occurrence is beyond the semester end, skip this session
                continue

            start_datetime = datetime.combine(
                first_occurrence_date, datetime.strptime(start_time_str, "%H:%M").time()
            )
            end_datetime = datetime.combine(
                first_occurrence_date, datetime.strptime(end_time_str, "%H:%M").time()
            )
            if end_datetime <= start_datetime:
                # Skipping event due to invalid time range
                continue

            event = Event()
            event.add("summary", f"{cls['name']} - Section {cls['group']}")
            event.add("location", session.get("class_room", "TBA"))
            event.add("dtstart", start_datetime)
            event.add("dtend", end_datetime)
            event.add("rrule", {"freq": "weekly", "until": end_date})

            # Add organizer (instructor) if available
            if instructor and instructor != "TBA":
                event.add("organizer", instructor)

            cal.add_component(event)

    with open(filename, "wb") as f:
        f.write(cal.to_ical())

    return filename


# ----------------------------

st.set_page_config(
    page_title="Nick's SAIT Schedule Builder",
    page_icon="📅",
)


# Check for authentication on app startup
def check_authentication():
    """Check if user has valid Banner API authentication tokens"""
    if "auth_checked" not in st.session_state:
        st.session_state.auth_checked = False

    if "banner_credentials" not in st.session_state:
        st.session_state.banner_credentials = None

    return st.session_state.auth_checked


def authentication_screen():
    """Display authentication screen for user to paste request headers"""
    st.title("🔐 Banner API Authentication")
    st.markdown("---")

    st.markdown(
        """
### Welcome to Nick's SAIT Schedule Builder!

To use the API features of this app (automatic course search and import), you need to provide your **Banner authentication tokens**.

---

#### 🔧 How to get your tokens

1. Open [**SAIT Banner Class Search**](https://sait-sust-prd-prd1-ban-ss-ssag6.sait.ca/StudentRegistrationSsb/ssb/classRegistration/classRegistration) in your browser.  
2. Press **F12** to open **Chrome DevTools**.  
3. Go to the **Network** tab.  
4. Make sure "Preserve log" is **checked** (top-left of the Network panel).  
5. Click **"Register for Courses"**, then select a **term** when prompted.  
6. In the course search box, type something simple (like `CPRG` or `ITSC`).  
7. In the Network tab, find a request called **`get_subjectcoursecombo?`**.  
8. Click it, then switch to the **Headers** tab on the right.  
9. Right-click anywhere in the request headers list → **Copy** → **Copy Request Headers**.  
10. Paste those headers into the input field below.

---

> 💡 **Tip:** You only need to do this once per login session.  
> If your session expires, just repeat the steps above to refresh your tokens.
"""
    )

    st.markdown("---")

    headers_input = st.text_area(
        "Paste your request headers here:",
        height=400,
        placeholder="""GET /StudentRegistrationSsb/ssb/classSearch/get_subjectcoursecombo?searchTerm=&term=202530 HTTP/1.1
Accept: application/json, text/javascript, */*; q=0.01
Cookie: JSESSIONID=...; NLB=...; NSC_ESNS=...
Host: sait-sust-prd-prd1-ban-ss-ssag6.sait.ca
X-Synchronizer-Token: ...
...""",
    )

    col1, col2 = st.columns([1, 3])

    with col1:
        if st.button("🚀 Parse & Continue", type="primary", use_container_width=True):
            if headers_input.strip():
                with st.spinner("Parsing headers..."):
                    result = parse_request_headers(headers_input)

                    if result["success"]:
                        # Format for Banner API
                        credentials = format_for_banner_api(result)

                        # Store in session state AND update the banner credentials for API calls
                        st.session_state.banner_credentials = credentials

                        # Update the Banner API cookies and token
                        cookie_dict = {}
                        for cookie in credentials["cookies"].split("; "):
                            if "=" in cookie:
                                name, value = cookie.split("=", 1)
                                cookie_dict[name] = value

                        st.session_state.banner_cookies = cookie_dict
                        st.session_state.banner_token = credentials["sync_token"]
                        st.session_state.auth_checked = True
                        st.session_state.term_selected = (
                            False  # Add term selection step
                        )

                        st.success("✅ Authentication successful!")
                        st.rerun()
                    else:
                        st.error("❌ Failed to parse headers")
                        for error in result["errors"]:
                            st.error(f"  - {error}")
                        st.info(
                            "💡 Make sure you copied the **full request headers** including Cookie and X-Synchronizer-Token"
                        )
            else:
                st.warning("⚠️ Please paste your request headers above")

    with col2:
        if st.button(
            "Skip Authentication (Limited Features)", use_container_width=True
        ):
            st.session_state.auth_checked = True
            st.session_state.term_selected = True  # Skip term selection too
            st.session_state.banner_credentials = None
            st.session_state.selected_term = "202530"  # Default term
            st.warning(
                "⚠️ You can use the app without authentication, but API features will be disabled"
            )
            st.rerun()

    st.divider()
    st.markdown("### ℹ️ Why do I need this?")
    st.markdown(
        """
    This app can automatically fetch course information from SAIT's Banner system, 
    but it requires your authentication tokens to make those requests on your behalf.
    
    **What we extract:**
    - Session cookies (JSESSIONID, NLB, NSC_ESNS)
    - Synchronizer token (for CSRF protection)
    
    **Privacy Note:**
    - All data stays in your browser session
        - Nothing is stored on our servers
        - Tokens expire after your Banner session ends
        
        **Alternative:**
        You can skip authentication and manually enter class information instead.
        
        **Don't trust me and want to read the code yourself?**  
        [https://github.com/SaladStik/NicksSaitScheduleBuilderPub](https://github.com/SaladStik/NicksSaitScheduleBuilderPub)  
        Here you go!
        """
    )


def term_selection_screen():
    """Display term selection screen"""
    st.title("📅 Select Term")
    st.markdown("---")

    st.markdown(
        """
    ### Choose the term/semester you want to build a schedule for
    """
    )

    with st.spinner("Fetching available terms..."):
        terms = fetch_available_terms()

    if not terms:
        st.error("❌ Failed to fetch terms. Using default Winter 2026 term.")
        if st.button("Continue with Default Term (Winter 2026)"):
            st.session_state.selected_term = "202530"
            st.session_state.term_selected = True
            st.rerun()
        return

    st.success(f"✅ Found {len(terms)} available terms")

    # Create options for selectbox
    term_options = {
        f"{term['description']} ({term['code']})": term["code"] for term in terms
    }

    selected_display = st.selectbox(
        "Select a term:",
        options=list(term_options.keys()),
        index=(
            1 if len(terms) > 1 else 0
        ),  # Default to second option (usually current term)
    )

    selected_code = term_options[selected_display]

    st.info(f"📚 Selected: **{selected_display}**")

    col1, col2 = st.columns([1, 3])

    with col1:
        if st.button("✅ Confirm Selection", type="primary", use_container_width=True):
            st.session_state.selected_term = selected_code
            st.session_state.term_selected = True
            st.success(f"Term set to: {selected_display}")
            st.rerun()

    with col2:
        if st.button("🔙 Back to Authentication"):
            st.session_state.auth_checked = False
            st.session_state.term_selected = False
            st.rerun()


# Main app function
def main():
    # Show welcome dialog first (before anything else)
    if "welcome_seen" not in st.session_state:
        st.session_state.welcome_seen = False

    if not st.session_state.welcome_seen:
        st.markdown(
            """
        <style>
        .welcome-dialog {
            background-color: #1e1e1e;
            padding: 2rem;
            border-radius: 10px;
            border: 2px solid #008CBA;
            margin: 2rem auto;
            max-width: 600px;
        }
        .welcome-title {
            color: #FFD700;
            font-size: 2rem;
            font-weight: bold;
            margin-bottom: 1rem;
            text-align: center;
        }
        .welcome-text {
            font-size: 1.1rem;
            line-height: 1.6;
            margin-bottom: 1.5rem;
            color: #ffffff;
        }
        .welcome-text p {
            color: #ffffff;
            margin: 1rem 0;
        }
        .welcome-text strong {
            color: #FFD700;
        }
        .social-links {
            margin-top: 1.5rem;
            padding: 1rem;
            background-color: #2d2d2d;
            border-radius: 5px;
        }
        .social-links p {
            color: #FFD700;
            font-weight: bold;
        }
        .social-link {
            display: block;
            margin: 0.5rem 0;
            color: #008CBA;
            text-decoration: none;
            font-weight: bold;
        }
        .social-link:hover {
            color: #00bfff;
        }
        </style>
        """,
            unsafe_allow_html=True,
        )

        st.markdown(
            """
        <div class="welcome-dialog">
            <div class="welcome-title">👋 Welcome to Nick's SAIT Schedule Builder!</div>
            <div class="welcome-text">
                <p>Hey there! Thanks for checking out my schedule builder app.</p>
                <p><strong>🎓 Made for SAIT students, by a SAIT student!</strong></p>
                <p><strong>📱 Mobile Notice:</strong> I'm not sure if this app will work perfectly on mobile devices. 
                If you run into issues, please try using a desktop or laptop browser. Sorry for any inconvenience!</p>
                <p><strong>💻 Technical Note:</strong> This app is biased towards technical users, but I did my best to make it usable for the general public. 
                If something seems confusing, don't hesitate to reach out!</p>
                <div class="social-links">
                    <p><strong>Connect with me:</strong></p>
                    <a href="https://www.linkedin.com/in/nicholas-irvine-303ab5284/" target="_blank" class="social-link">
                        💼 LinkedIn - Nicholas Irvine
                    </a>
                    <a href="https://github.com/SaladStik" target="_blank" class="social-link">
                        🐙 GitHub - @SaladStik
                    </a>
                </div>
            </div>
        </div>
        """,
            unsafe_allow_html=True,
        )

        st.markdown("---")

        col1, col2, col3 = st.columns([1, 1, 1])
        with col2:
            if st.button(
                "✅ Got it, Let's Go!", type="primary", use_container_width=True
            ):
                st.session_state.welcome_seen = True
                st.rerun()

        return

    # Check authentication first
    if not check_authentication():
        authentication_screen()
        return

    # Check if term is selected
    if not st.session_state.get("term_selected", False):
        term_selection_screen()
        return

    st.title("Nick's SAIT Schedule Builder 🎓")
    st.caption("Browse through ALL possible schedule combinations!")

    # Sidebar for authentication status
    st.sidebar.header("🔐 API Authentication")

    if st.session_state.banner_credentials:
        st.sidebar.success("✅ Authenticated")

        # Show selected term
        selected_term = st.session_state.get("selected_term", "202530")
        st.sidebar.info(f"📅 Term: {selected_term}")

        if st.sidebar.button("🔄 Change Term"):
            st.session_state.term_selected = False
            st.rerun()

        if st.sidebar.button("🔄 Update Tokens"):
            st.session_state.auth_checked = False
            st.session_state.term_selected = False
            st.rerun()

        # Show token info (truncated)
        with st.sidebar.expander("View Token Info"):
            if "sync_token" in st.session_state.banner_credentials:
                token = st.session_state.banner_credentials["sync_token"]
                st.sidebar.code(f"Sync Token: {token[:20]}...", language=None)
            if "cookies" in st.session_state.banner_credentials:
                cookies = st.session_state.banner_credentials["cookies"]
                st.sidebar.code(f"Cookies: {cookies[:50]}...", language=None)
    else:
        st.sidebar.warning("⚠️ Not authenticated")
        st.sidebar.info("API features disabled")
        if st.sidebar.button("🔐 Authenticate Now"):
            st.session_state.auth_checked = False
            st.rerun()

    st.sidebar.markdown("---")

    # Info banner about stateless storage
    st.info(
        "💡 **Stateless App**: All data is stored in your browser session only. Data will be lost when you refresh or close the tab. Download your schedules to save them!",
        icon="ℹ️",
    )

    # Use a default ID for all users (stateless - no user accounts needed)
    id = "default_user"
    st.session_state.id = id

    # Main app content
    st.markdown(
        """
        <style>
        html, body, [class*="st-"] {
            font-family: 'Gill Sans', 'Gill Sans MT', Calibri, 'Trebuchet MS', sans-serif;
        }
        .stButton>button {
            border: none;
            border-radius: px;
            padding: 10px 24px;
            margin: 5px 0px;
            color: white;
            background-color: #008CBA;
        }
        .stButton>button:hover {
            background-color: #005f73;
        }
        </style>
    """,
        unsafe_allow_html=True,
    )

    st.markdown("Created by [SaladStik](https://github.com/SaladStik)")

    st.divider()
    st.markdown("### 📖 How to use the app?")
    st.markdown(
        """
        <div >
            <ol>
                <li><strong>Register your classes:</strong> Start by entering all the classes you're interested in. Whether you're attending them or considering them, to see all possible schedules.</li>
                <li><strong>Create possible schedules:</strong> Navigate to the 'Schedule Creator' tab to generate customized schedules based on your classes of interest. The app will show you ALL possible combinations and let you browse through them!</li>
                <li><strong>Download the ICS file:</strong> After finishing choosing the classes to attend, go to the 'Add to calendar' section. There, you can download an ICS file, which can be easily added to your calendar.</li>
            </ol>
        </div>
    """,
        unsafe_allow_html=True,
    )
    st.divider()

    # Create tabs for different functionalities
    tab1, tab2, tab3, tab4 = st.tabs(
        [
            "📚 Class Registration",
            "⏱️ Schedule Creator",
            "🗑️ Drop Classes",
            "📅 Add to Calendar",
        ]
    )

    with tab1:
        class_logger()

    with tab2:
        timetable_creator()

    with tab3:
        manual_schedule_editor()

    with tab4:
        calendar_ics_generator()


def remove_class_from_db(id, class_name, group):
    """Remove class from browser session state (stateless)"""
    if id in st.session_state.classes_data:
        st.session_state.classes_data[id] = [
            cls
            for cls in st.session_state.classes_data[id]
            if not (cls["name"] == class_name and cls["group"] == group)
        ]


# Streamlit app function
def class_logger():
    id = st.session_state.get("id", "default_user")

    tab1, tab2, tab3, tab4 = st.tabs(
        [
            "🔍 Search & Add Classes",
            "Add class manually",
            "🌐 Fetch from API",
            "Delete classes",
        ]
    )

    # NEW TAB 1: Search and Add Classes
    with tab1:
        st.subheader("🔍 Search for Courses")
        st.markdown("Search for any course and add available sections to your schedule")

        # Filter checkbox
        filter_available_only = st.checkbox(
            "✅ Filter by available seats only (leave this on for 99% of cases)",
            value=True,
            key="filter_available_only",
        )

        # Search bar
        col1, col2 = st.columns([4, 1])
        with col1:
            search_term = st.text_input(
                "Course",
                placeholder="Type course code or name (e.g., ABDY, ITSC, Computer)",
                label_visibility="collapsed",
                key="course_search",
            )
        with col2:
            search_button = st.button(
                "🔍 Search", type="primary", use_container_width=True
            )

        # Perform search when user types or clicks search
        if search_term and len(search_term) >= 2:
            if (
                "last_search" not in st.session_state
                or st.session_state.last_search != search_term
                or search_button
            ):
                st.session_state.last_search = search_term
                selected_term = st.session_state.get("selected_term", "202530")
                st.session_state.search_results = search_courses(
                    search_term, term=selected_term
                )

            if st.session_state.get("search_results"):
                results = st.session_state.search_results

                # Create dropdown options
                course_options = [f"{r['code']} {r['description']}" for r in results]

                st.markdown(f"### Available Seats")
                st.markdown(f"Found {len(results)} course(s)")

                # Dropdown selection (styled like the screenshot)
                selected_course = st.selectbox(
                    "Select a course to view sections:",
                    options=course_options,
                    label_visibility="collapsed",
                )

                if selected_course:
                    # Extract course code from selection
                    selected_code = selected_course.split()[0]

                    # Button to fetch sections for selected course
                    if st.button(
                        f"📚 Load Sections for {selected_code}",
                        type="primary",
                        use_container_width=True,
                    ):
                        # Reset search and fetch sections
                        selected_term = st.session_state.get("selected_term", "202530")
                        reset_banner_search(selected_term)

                        # Get filter setting
                        open_only = st.session_state.get("filter_available_only", True)

                        with st.spinner(f"Fetching sections for {selected_code}..."):
                            import time

                            session_id = f"streamlit{int(time.time() * 1000)}"
                            api_response = fetch_banner_api(
                                selected_term,
                                selected_code,
                                session_id,
                                open_only=open_only,
                            )

                            if api_response:
                                classes = parse_banner_response(
                                    api_response, open_only=open_only
                                )

                                if classes:
                                    # Store in session state so it persists
                                    st.session_state.fetched_classes = classes
                                    st.session_state.fetched_course_code = selected_code

                    # Display selection interface if we have fetched classes
                    if (
                        st.session_state.get("fetched_classes")
                        and st.session_state.get("fetched_course_code") == selected_code
                    ):
                        classes = st.session_state.fetched_classes

                        # Check if we have 3-character sections with course prefixes
                        section_prefixes = {}
                        has_course_prefixes = False

                        for cls in classes:
                            section = cls["group"]
                            if len(section) == 3:
                                # Extract first 2 characters as course prefix
                                prefix = section[:2]
                                if prefix not in section_prefixes:
                                    section_prefixes[prefix] = []
                                section_prefixes[prefix].append(cls)
                                has_course_prefixes = True
                            else:
                                # Regular sections without course prefix
                                if "ALL" not in section_prefixes:
                                    section_prefixes["ALL"] = []
                                section_prefixes["ALL"].append(cls)

                        # If we detected course prefixes, let user select which ones to import
                        if has_course_prefixes and len(section_prefixes) > 1:
                            st.success(
                                f"✅ Found {len(classes)} section(s) across {len(section_prefixes)} course(s)"
                            )
                            st.info(
                                "📋 **Multiple course sections detected!** Select which sections to import:"
                            )

                            # Create options for multiselect - show course prefix with section details
                            prefix_options = []
                            for prefix in sorted(section_prefixes.keys()):
                                count = len(section_prefixes[prefix])
                                # Get sample section names
                                sample_sections = ", ".join(
                                    [
                                        cls["group"]
                                        for cls in section_prefixes[prefix][:3]
                                    ]
                                )
                                if count > 3:
                                    sample_sections += f" (+{count-3} more)"

                                prefix_options.append(
                                    f"{prefix} - {count} section(s) ({sample_sections})"
                                )

                            # Multiselect for course prefixes
                            selected_prefix_labels = st.multiselect(
                                "Select course sections to import:",
                                prefix_options,
                                placeholder="Choose sections",
                                key="section_multiselect",
                            )

                            # Import button for selected prefixes
                            if selected_prefix_labels:
                                # Extract prefix codes from labels
                                selected_prefixes = [
                                    label.split(" - ")[0]
                                    for label in selected_prefix_labels
                                ]

                                # Show preview of what will be imported
                                classes_to_import = []
                                for prefix in selected_prefixes:
                                    classes_to_import.extend(section_prefixes[prefix])

                                st.markdown(
                                    f"**📋 Ready to import {len(classes_to_import)} section(s):**"
                                )
                                for cls in classes_to_import:
                                    st.text(
                                        f"  • {cls['name']}-{cls['group']} ({cls['seats_available']} seats)"
                                    )

                                if st.button(
                                    "✨ Import Selected Sections", type="primary"
                                ):
                                    # Auto-import selected classes
                                    imported_count = 0
                                    for cls in classes_to_import:
                                        try:
                                            save_class_to_db(
                                                id,
                                                {
                                                    "name": cls["name"],
                                                    "group": cls["group"],
                                                    "schedule": cls["schedule"],
                                                    "crn": cls.get("crn", ""),
                                                    "instructor": cls.get(
                                                        "instructor", "TBA"
                                                    ),
                                                    "seats_available": cls.get(
                                                        "seats_available", 0
                                                    ),
                                                },
                                            )
                                            imported_count += 1
                                        except Exception as e:
                                            st.error(
                                                f"Error importing {cls['name']}-{cls['group']}: {e}"
                                            )

                                    if imported_count > 0:
                                        st.success(
                                            f"✨ Successfully added {imported_count} section(s) to your schedule!"
                                        )
                                        st.balloons()

                                        # Show what was added
                                        st.markdown("### 📚 Classes Added:")
                                        for cls in classes_to_import:
                                            st.markdown(
                                                f"#### {cls['name']}-{cls['group']} ({cls['seats_available']} seats available)"
                                            )
                                            st.write(
                                                f"**Instructor:** {cls.get('instructor', 'TBA')}"
                                            )
                                            st.write(f"**CRN:** {cls.get('crn', '')}")
                                            st.write(
                                                f"**Seats:** {cls['seats_available']} / {cls.get('max_enrollment', 'N/A')}"
                                            )
                                            st.write("**Schedule:**")
                                            for session in cls["schedule"]:
                                                st.write(
                                                    f"  • {session['day']}: {session['start_time']}-{session['end_time']} @ {session['class_room']}"
                                                )
                                            st.divider()

                                        st.info(
                                            "👉 Go to **Schedule Creator** to see possible schedules!"
                                        )

                                        # Clear session state after import
                                        del st.session_state.fetched_classes
                                        del st.session_state.fetched_course_code
                            else:
                                st.info("👆 Select sections from the dropdown above")

                        else:
                            # No course prefixes detected, import all as usual
                            st.success(
                                f"✅ Found {len(classes)} available section(s) for {selected_code}"
                            )

                            # Auto-import
                            imported_count = 0
                            for cls in classes:
                                try:
                                    save_class_to_db(
                                        id,
                                        {
                                            "name": cls["name"],
                                            "group": cls["group"],
                                            "schedule": cls["schedule"],
                                            "crn": cls.get("crn", ""),
                                            "instructor": cls.get("instructor", "TBA"),
                                            "seats_available": cls.get(
                                                "seats_available", 0
                                            ),
                                        },
                                    )
                                    imported_count += 1
                                except Exception as e:
                                    st.error(
                                        f"Error importing {cls['name']}-{cls['group']}: {e}"
                                    )

                            if imported_count > 0:
                                st.success(
                                    f"✨ Automatically added {imported_count} section(s) to your schedule!"
                                )
                                st.balloons()

                                # Show what was added
                                st.markdown("### 📚 Classes Added:")
                                for cls in classes:
                                    st.markdown(
                                        f"#### {cls['name']}-{cls['group']} ({cls['seats_available']} seats available)"
                                    )
                                    st.write(
                                        f"**Instructor:** {cls.get('instructor', 'TBA')}"
                                    )
                                    st.write(f"**CRN:** {cls.get('crn', '')}")
                                    st.write(
                                        f"**Seats:** {cls['seats_available']} / {cls.get('max_enrollment', 'N/A')}"
                                    )
                                    st.write("**Schedule:**")
                                    for session in cls["schedule"]:
                                        st.write(
                                            f"  • {session['day']}: {session['start_time']}-{session['end_time']} @ {session['class_room']}"
                                        )
                                    st.divider()

                                st.info(
                                    "👉 Go to **Schedule Creator** to see possible schedules!"
                                )

                                # Clear session state after import
                                del st.session_state.fetched_classes
                                del st.session_state.fetched_course_code
            else:
                st.info("No courses found. Try a different search term.")
        elif search_term:
            st.info("Type at least 2 characters to search")
        else:
            st.info("👆 Enter a course code or name above to search")

    # TAB 2: Manual entry
    with tab2:
        class_name = st.text_input("Class name")
        group_section = st.text_input("Group/Section")

        # Input for number of days and dynamic schedule inputs
        num_days = st.number_input(
            "Number of days per week", min_value=1, max_value=10, step=1, key="num_days"
        )
        schedule_entries = []
        for i in range(num_days):
            cols = st.columns(4)
            with cols[0]:
                day = st.selectbox(
                    f"Day {i+1}",
                    ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"],
                    key=f"day{i}",
                )
            with cols[1]:
                start_time = st.time_input(
                    f"Start time {i+1}", key=f"start_time{i}", value=None
                )
            with cols[2]:
                end_time = st.time_input(
                    f"End time {i+1}", key=f"end_time{i}", value=None
                )
            with cols[3]:
                class_room = st.text_input(f"Room {i+1}", key=f"class_room{i}")
            schedule_entries.append((day, start_time, end_time, class_room))

        # Submission button
        submit_button = st.button("Register class")

        # Handling the submission
        if submit_button and class_name and group_section:
            schedule = [
                {
                    "day": day,
                    "start_time": start_time.strftime("%H:%M"),
                    "end_time": end_time.strftime("%H:%M"),
                    "class_room": class_room,
                }
                for day, start_time, end_time, class_room in schedule_entries
            ]

            new_class = {
                "name": class_name.capitalize(),
                "group": group_section,
                "schedule": schedule,
            }

            save_class_to_db(id, new_class)

            display_class_entry(new_class)
            st.success("Class registered and saved successfully.")

    # TAB 3: Fetch from API
    with tab3:
        st.subheader("🌐 Fetch Classes Directly from Banner API")

        # ========== FETCH ALL FROM API ==========
        st.markdown("### 🚀 Fetch Fresh Data from SAIT Banner")
        if st.button("🚀 Load ALL Classes from Files (classes folder)", type="primary"):
            import os
            import json

            classes_folder = "classes"
            all_fetched_classes = []

            if os.path.exists(classes_folder):
                st.info(f"Reading class files from {classes_folder} folder...")

                # Read all class files
                for filename in os.listdir(classes_folder):
                    if filename.endswith(".txt"):
                        filepath = os.path.join(classes_folder, filename)
                        try:
                            with open(filepath, "r", encoding="utf-8") as f:
                                data = json.load(f)
                                classes = parse_banner_response(data)
                                all_fetched_classes.extend(classes)
                                st.success(
                                    f"✅ Loaded {len(classes)} sections from {filename}"
                                )
                        except Exception as e:
                            st.warning(f"⚠️ Could not parse {filename}: {e}")

                if all_fetched_classes:
                    st.success(
                        f"🎉 Successfully loaded {len(all_fetched_classes)} class section(s) with available seats!"
                    )

                    # Display preview
                    st.markdown("### Preview of Loaded Classes")
                    for cls in all_fetched_classes:
                        st.markdown(
                            f"#### 📚 {cls['name']} - Section {cls['group']} ({cls['seats_available']} seats available)"
                        )
                        st.write(f"**Instructor:** {cls.get('instructor', 'TBA')}")
                        st.write(f"**CRN:** {cls.get('crn', 'N/A')}")
                        st.write(
                            f"**Seats:** {cls['seats_available']} / {cls.get('max_enrollment', 'N/A')}"
                        )
                        st.write("**Schedule:**")
                        for session in cls["schedule"]:
                            st.write(
                                f"  • {session['day']}: {session['start_time']} - {session['end_time']} ({session['class_room']})"
                            )
                        st.divider()

                    # Import button
                    if st.button(
                        "✅ Import All Loaded Classes to My Schedule", key="file_import"
                    ):
                        for cls in all_fetched_classes:
                            # Remove extra metadata before saving
                            save_cls = {
                                "name": cls["name"],
                                "group": cls["group"],
                                "schedule": cls["schedule"],
                                "crn": cls.get("crn", ""),
                                "instructor": cls.get("instructor", "TBA"),
                                "seats_available": cls.get("seats_available", 0),
                            }
                            save_class_to_db(id, save_cls)
                        st.success(
                            f"✨ Successfully imported {len(all_fetched_classes)} classes to your schedule!"
                        )
                        st.balloons()
                else:
                    st.warning("No classes with available seats found in the files.")
            else:
                st.error(
                    f"Classes folder not found at: {os.path.abspath(classes_folder)}"
                )

        st.markdown("---")
        st.markdown("### 📋 Paste API Response")
        st.markdown("Copy the JSON response from Banner API and paste it below")

        api_json_input = st.text_area(
            "Paste Banner API JSON Response",
            height=300,
            placeholder='{"success": true, "data": [...]}',
            help="Paste the entire JSON response from the Banner API network call",
        )

        if st.button("� Load from Pasted JSON", type="secondary"):
            if api_json_input.strip():
                try:
                    data = json.loads(api_json_input)
                    classes = parse_banner_response(data)

                    if classes:
                        st.success(f"✅ Parsed {len(classes)} class section(s)!")

                        # Display preview
                        st.markdown("### Preview")
                        for cls in classes:
                            st.markdown(
                                f"#### 📚 {cls['name']} - Section {cls['group']} ({cls['seats_available']} seats)"
                            )
                            st.write(f"**Instructor:** {cls.get('instructor', 'TBA')}")
                            st.write(f"**Schedule:**")
                            for session in cls["schedule"]:
                                st.write(
                                    f"  • {session['day']}: {session['start_time']} - {session['end_time']} ({session['class_room']})"
                                )
                            st.divider()

                        # Import button
                        if st.button("✅ Import These Classes", key="paste_import"):
                            for cls in classes:
                                save_cls = {
                                    "name": cls["name"],
                                    "group": cls["group"],
                                    "schedule": cls["schedule"],
                                    "crn": cls.get("crn", ""),
                                    "instructor": cls.get("instructor", "TBA"),
                                    "seats_available": cls.get("seats_available", 0),
                                }
                                save_class_to_db(id, save_cls)
                            st.success(f"✨ Imported {len(classes)} classes!")
                    else:
                        st.warning("No classes with available seats found in the JSON")
                except json.JSONDecodeError as e:
                    st.error(f"Invalid JSON: {e}")
            else:
                st.warning("Please paste JSON data first")

        # NEW: Direct API Fetch Button
        st.markdown("---")
        st.markdown("### 🚀 Fetch Fresh Data Directly from SAIT Banner API")

        col1, col2 = st.columns([3, 1])
        with col1:
            st.info(
                "Fetch the latest class data with available seats from all 6 courses"
            )
        with col2:
            if st.button("🔑 Update Auth", help="Update cookies/token if expired"):
                st.session_state.show_auth = True

        # Show auth form if needed
        if st.session_state.get("show_auth", False):
            st.markdown("### 🔐 Update Banner API Credentials")
            st.markdown(
                """
            Get these from Chrome DevTools:
            1. Open SAIT Banner class search  
            2. Press F12 → Network tab
            3. Open the 'Register for Courses' page
            4. Select A Term
            5. Search for a class
            6. Find `get_subjectcoursecombo?` and select it + right click it
            7. Copy values from Headers tab
            """
            )

            jsessionid = st.text_input(
                "JSESSIONID",
                value=st.session_state.get("banner_cookies", {}).get("JSESSIONID", ""),
            )
            nlb = st.text_input(
                "NLB", value=st.session_state.get("banner_cookies", {}).get("NLB", "")
            )
            nsc = st.text_input(
                "NSC_ESNS",
                value=st.session_state.get("banner_cookies", {}).get("NSC_ESNS", ""),
            )
            sync_token = st.text_input(
                "X-Synchronizer-Token", value=st.session_state.get("banner_token", "")
            )

            if st.button("💾 Save"):
                st.session_state.banner_cookies = {
                    "JSESSIONID": jsessionid,
                    "NLB": nlb,
                    "NSC_ESNS": nsc,
                }
                st.session_state.banner_token = sync_token
                st.session_state.show_auth = False
                st.success("✅ Saved!")
                st.rerun()

        # Big fetch button
        if st.button(
            "🔄 FETCH ALL CLASSES FROM API NOW",
            type="primary",
            use_container_width=True,
        ):
            course_codes = [
                "ITSC320",
                "CPSY300",
                "CPSY301",
                "INTP302",
                "CPRG303",
                "CPRG305",
            ]
            selected_term = st.session_state.get("selected_term", "202530")
            term = selected_term

            st.info(f"Fetching: {', '.join(course_codes)} for term {term}")
            all_fetched_classes = fetch_all_available_courses(
                term, course_codes, open_only=True
            )

            if all_fetched_classes:
                st.success(f"🎉 Fetched {len(all_fetched_classes)} sections!")

                # AUTO-IMPORT: Save all classes immediately
                imported_count = 0
                for cls in all_fetched_classes:
                    try:
                        save_class_to_db(
                            id,
                            {
                                "name": cls["name"],
                                "group": cls["group"],
                                "schedule": cls["schedule"],
                                "crn": cls.get("crn", ""),
                                "instructor": cls.get("instructor", "TBA"),
                                "seats_available": cls.get("seats_available", 0),
                            },
                        )
                        imported_count += 1
                    except Exception as e:
                        st.error(f"Error importing {cls['name']}-{cls['group']}: {e}")

                if imported_count > 0:
                    st.success(
                        f"✨ Automatically imported {imported_count} classes to your schedule!"
                    )
                    st.balloons()

                    # Show what was imported
                    st.markdown("### 📚 Classes Added:")
                    for cls in all_fetched_classes:
                        st.markdown(
                            f"#### {cls['name']}-{cls['group']} ({cls['seats_available']} seats)"
                        )
                        st.write(f"**Instructor:** {cls.get('instructor', 'TBA')}")
                        st.write(f"**CRN:** {cls.get('crn', '')}")
                        for session in cls["schedule"]:
                            st.write(
                                f"{session['day']}: {session['start_time']}-{session['end_time']} @ {session['class_room']}"
                            )
                        st.divider()

                    st.info(
                        "👉 Go to the **Schedule Creator** tab to see possible schedules!"
                    )
                else:
                    st.error("Failed to import any classes")
            else:
                st.warning("⚠️ No classes found. Check your auth credentials.")

    # TAB 4: Delete classes
    with tab4:
        classes = get_classes_from_db(id)
        if classes:
            class_names = [f"{cls['name']} - {cls['group']}" for cls in classes]
            selected_classes = st.multiselect(
                "Select classes you want to delete", class_names
            )
            if st.button("Delete classes"):
                for selected_class in selected_classes:
                    class_name, group = selected_class.split(" - ")
                    remove_class_from_db(id, class_name, group)
                st.success("Classes deleted successfully")
        else:
            st.warning("You don't have any registered classes to delete")


# Timetable Creator tab - ENHANCED VERSION
def timetable_creator():
    id = st.session_state.get("id")
    if id:
        classes = get_classes_from_db(id)
        if not classes:
            st.warning(
                "No classes found for the entered ID. Please register some classes first."
            )
            return

        parsed_classes = [
            {
                **cls,
                "schedule": [
                    {
                        **session,
                        "start_time": parse_time(session["start_time"]),
                        "end_time": parse_time(session["end_time"]),
                    }
                    for session in cls["schedule"]
                ],
            }
            for cls in classes
        ]

        # Get unique class names
        class_names = list(set(cls["name"] for cls in parsed_classes))

        # Add multiselect for mandatory classes
        mandatory_classes = st.multiselect(
            "Select classes you'd like to include in possible schedules",
            class_names,
            placeholder="Choose classes",
            max_selections=6,
        )

        # User input for selecting free days
        days_of_week = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
        free_days = st.multiselect(
            "Select days you don't want to have classes",
            days_of_week,
            placeholder="Choose days",
        )

        # Add option to search for specific combination
        st.markdown("---")
        st.markdown("**🎯 Test a Specific Schedule Combination**")
        st.caption(
            "Enter the exact sections you want to check (e.g., your actual schedule from your student portal)"
        )

        test_combo_text = st.text_area(
            "Enter class sections (one per line): ClassName Section",
            placeholder="INTP 302 B\nCPRG 303 B\nCPSY 301 A\nITSC 320 A\nCPRG 305 D",
            height=100,
        )

        if st.button("🔎 Check This Specific Combination"):
            if test_combo_text.strip():
                lines = [
                    line.strip()
                    for line in test_combo_text.strip().split("\n")
                    if line.strip()
                ]
                test_classes = []

                for line in lines:
                    parts = line.rsplit(
                        " ", 1
                    )  # Split from the right to handle multi-word class names
                    if len(parts) == 2:
                        class_name, section = parts
                        # Find this class in parsed_classes
                        matching_class = next(
                            (
                                cls
                                for cls in parsed_classes
                                if cls["name"] == class_name.strip()
                                and cls["group"] == section.strip()
                            ),
                            None,
                        )
                        if matching_class:
                            test_classes.append(matching_class)
                        else:
                            st.warning(
                                f"⚠️ Could not find: {class_name} Section {section}"
                            )

                if test_classes:
                    st.info(f"Found {len(test_classes)} classes to test")

                    # Check for conflicts in this specific combination
                    conflicts_found = []
                    for cls1, cls2 in combinations(test_classes, 2):
                        if has_conflict(cls1, cls2):
                            conflicts_found.append(
                                f"{cls1['name']}({cls1['group']}) ⚔️ {cls2['name']}({cls2['group']})"
                            )

                            # Show detailed conflict info
                            for s1 in cls1["schedule"]:
                                for s2 in cls2["schedule"]:
                                    if s1["day"] == s2["day"]:
                                        if times_overlap(
                                            s1["start_time"],
                                            s1["end_time"],
                                            s2["start_time"],
                                            s2["end_time"],
                                        ):
                                            st.error(
                                                f"❌ {cls1['name']}({cls1['group']}) and {cls2['name']}({cls2['group']}) both have {s1['day']}: {s1['start_time']}-{s1['end_time']} vs {s2['start_time']}-{s2['end_time']}"
                                            )

                    if conflicts_found:
                        st.error(
                            f"❌ This combination has {len(conflicts_found)} conflict(s)"
                        )
                        for conflict in conflicts_found:
                            st.text(f"⚠️ {conflict}")
                    else:
                        st.success(
                            "✅ This combination has NO conflicts! It's a valid schedule."
                        )
                        st.balloons()

        st.markdown("---")

        generate_button = st.button("Generate ALL Schedule Options")

        if generate_button:
            try:
                # Filter classes based on selection
                if mandatory_classes:
                    # Only use classes that were selected
                    filtered_classes = [
                        cls
                        for cls in parsed_classes
                        if cls["name"] in mandatory_classes
                    ]
                else:
                    # If nothing selected, use all classes
                    filtered_classes = parsed_classes

                if not filtered_classes:
                    st.warning(
                        "⚠️ Please select at least one class to generate schedules"
                    )
                    return

                # Generate ALL possible combinations - with priority scoring
                all_combinations_with_score = []
                excluded_count = 0
                conflict_details = []

                # Try different class counts from maximum down to minimum
                max_classes = min(6, len(filtered_classes))
                min_classes = 1  # Always start from 1 to show all options

                for num_classes in range(max_classes, min_classes - 1, -1):
                    class_combinations = combinations(filtered_classes, num_classes)

                    for combo in class_combinations:
                        # Check for conflicts
                        has_time_conflict = False
                        conflict_info = []

                        for cls1, cls2 in combinations(combo, 2):
                            if has_conflict(cls1, cls2):
                                has_time_conflict = True
                                conflict_info.append(
                                    f"{cls1['name']}({cls1['group']}) conflicts with {cls2['name']}({cls2['group']})"
                                )

                        if has_time_conflict:
                            excluded_count += 1
                            conflict_details.extend(conflict_info)
                            continue

                        if not has_unique_classes(combo):
                            excluded_count += 1
                            continue

                        # Calculate a priority score based on preferences
                        score = 0

                        # Higher score for more classes
                        score += num_classes * 1000

                        # Bonus points if it respects free days
                        if has_free_days(combo, free_days):
                            score += 500

                        # Bonus points for each mandatory class included
                        mandatory_count = sum(
                            1 for cls in combo if cls["name"] in mandatory_classes
                        )
                        score += mandatory_count * 200

                        # Check if ALL mandatory classes are included
                        has_all_mandatory = all(
                            any(cls["name"] == mandatory_class for cls in combo)
                            for mandatory_class in mandatory_classes
                        )
                        if has_all_mandatory:
                            score += 300

                        all_combinations_with_score.append(
                            (
                                score,
                                num_classes,
                                combo,
                                has_all_mandatory,
                                has_free_days(combo, free_days),
                            )
                        )

                if not all_combinations_with_score:
                    st.error(
                        "❌ No valid schedules found. This might happen if all class combinations have time conflicts."
                    )
                    st.info(
                        "💡 Try adding more class sections with different time slots."
                    )
                    return

                # Sort by score (highest first) to show best options first
                all_combinations_with_score.sort(key=lambda x: x[0], reverse=True)

                # Store in session state
                st.session_state["all_schedule_options"] = all_combinations_with_score
                st.session_state["current_schedule_index"] = 0

                # Count perfect matches
                perfect_matches = sum(
                    1 for item in all_combinations_with_score if item[3] and item[4]
                )
                matches_with_mandatory = sum(
                    1 for item in all_combinations_with_score if item[3]
                )

                st.success(
                    f"✨ Found {len(all_combinations_with_score)} possible schedule combinations!"
                )
                st.info(
                    f"📊 Best schedules have {all_combinations_with_score[0][1]} classes"
                )

                if perfect_matches > 0:
                    st.success(
                        f"🎯 {perfect_matches} schedules match ALL your criteria (mandatory classes + free days)"
                    )
                elif matches_with_mandatory > 0:
                    st.warning(
                        f"⚠️ {matches_with_mandatory} schedules include all mandatory classes (but may not respect free days)"
                    )
                else:
                    st.warning(
                        "⚠️ No schedules include all mandatory classes. Showing best available options."
                    )

            except Exception as e:
                st.error(f"An error occurred: {e}")

        # Display schedules if they exist
        if (
            "all_schedule_options" in st.session_state
            and st.session_state["all_schedule_options"]
        ):
            all_combos = st.session_state["all_schedule_options"]
            current_idx = st.session_state.get("current_schedule_index", 0)

            st.markdown("---")
            st.subheader(f"📅 Browse Schedule Options ({len(all_combos)} total)")

            # Navigation controls
            col1, col2, col3, col4, col5 = st.columns([1, 1, 2, 1, 1])

            with col1:
                if st.button("⬅️ Previous", disabled=(current_idx == 0)):
                    st.session_state["current_schedule_index"] = max(0, current_idx - 1)
                    st.rerun()

            with col2:
                if st.button("➡️ Next", disabled=(current_idx >= len(all_combos) - 1)):
                    st.session_state["current_schedule_index"] = min(
                        len(all_combos) - 1, current_idx + 1
                    )
                    st.rerun()

            with col3:
                # Jump to specific schedule
                new_idx = (
                    st.number_input(
                        "Go to schedule #",
                        min_value=1,
                        max_value=len(all_combos),
                        value=current_idx + 1,
                        key="schedule_jumper",
                    )
                    - 1
                )
                if new_idx != current_idx:
                    st.session_state["current_schedule_index"] = new_idx
                    st.rerun()

            with col4:
                if st.button("⏮️ First"):
                    st.session_state["current_schedule_index"] = 0
                    st.rerun()

            with col5:
                if st.button("⏭️ Last"):
                    st.session_state["current_schedule_index"] = len(all_combos) - 1
                    st.rerun()

            # Display current schedule
            score, num_classes, current_combo, has_all_mandatory, respects_free_days = (
                all_combos[current_idx]
            )

            st.markdown(f"### Schedule Option #{current_idx + 1} of {len(all_combos)}")
            st.markdown(f"**Classes in this schedule:** {num_classes}")

            # Show match indicators
            col_status1, col_status2 = st.columns(2)
            with col_status1:
                if has_all_mandatory:
                    st.success("✅ Includes all mandatory classes")
                else:
                    st.warning("⚠️ Missing some mandatory classes")

            with col_status2:
                if respects_free_days:
                    st.success("✅ Respects your free days")
                else:
                    st.warning("⚠️ Uses some of your preferred free days")

            # Show class list
            class_list = ", ".join(
                [f"{cls['name']} ({cls['group']})" for cls in current_combo]
            )
            st.info(f"📚 {class_list}")

            # Show missing courses if any
            if mandatory_classes:
                courses_in_schedule = set(cls["name"] for cls in current_combo)
                missing_courses = set(mandatory_classes) - courses_in_schedule

                if missing_courses:
                    st.warning(
                        f"⚠️ **Missing courses:** {', '.join(sorted(missing_courses))}"
                    )
                else:
                    st.success("✅ All selected courses are included in this schedule")

            # Display the timetable
            time_slots = get_unique_time_slots(parsed_classes)
            timetable_html = display_timetable_html(current_combo, time_slots, classes)
            st.markdown(timetable_html, unsafe_allow_html=True)

            # Download options
            st.markdown("---")
            col_dl1, col_dl2, col_dl3 = st.columns(3)

            with col_dl1:
                # Download this specific schedule
                if st.button("💾 Download This Schedule (Excel)"):
                    filename = f"schedule_option_{current_idx + 1}.xlsx"
                    create_single_sheet_xlsx_timetables(
                        [current_combo], filename, time_slots, classes
                    )
                    with open(filename, "rb") as file:
                        st.download_button(
                            label="⬇️ Download Excel",
                            data=file,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )

            with col_dl2:
                # Refresh colors button
                if st.button("🎨 Switch Schedule Colors", use_container_width=True):
                    st.rerun()

            with col_dl3:
                # Download ALL schedules
                if st.button("📚 Download ALL Schedules (Excel)"):
                    filename = "all_schedule_options.xlsx"
                    all_combos_only = [combo for _, _, combo, _, _ in all_combos]
                    create_single_sheet_xlsx_timetables(
                        all_combos_only, filename, time_slots, classes
                    )
                    with open(filename, "rb") as file:
                        st.download_button(
                            label=f"⬇️ Download All {len(all_combos)} Schedules",
                            data=file,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )

            # Apply Schedule to Banner
            st.markdown("---")
            st.markdown("### 🎓 Add This Schedule to SAIT Banner")

            # Check if user is authenticated
            if not st.session_state.get("banner_credentials"):
                st.warning(
                    "⚠️ You need to authenticate first to use the auto-registration feature"
                )
                st.info(
                    "👆 Go to the sidebar and click 'Authenticate Now' to enable this feature"
                )
            else:
                st.info(
                    "ℹ️ This will add these classes to your registration cart and automatically submit them"
                )
                st.warning(
                    "⚠️ If you need to drop classes, please navigate to the Drop tab first"
                )

                # Show what will be registered
                st.markdown("### 📋 Classes that will be added to cart")
                for cls in current_combo:
                    crn = cls.get("crn", "N/A")
                    instructor = cls.get("instructor", "TBA")
                    seats = cls.get("seats_available", "?")
                    st.write(
                        f"• **{cls['name']}** - Section {cls['group']} (CRN: {crn})"
                    )
                    st.write(f"  - Instructor: {instructor}")
                    st.write(f"  - Seats Available: {seats}")
                    st.write("")
                st.divider()

                # Confirmation checkbox
                confirm_apply = st.checkbox(
                    "✅ I understand this will add these classes to my registration cart"
                )

                if st.button(
                    "🚀 ADD SCHEDULE TO BANNER CART",
                    type="primary",
                    disabled=not confirm_apply,
                    use_container_width=True,
                ):
                    if not confirm_apply:
                        st.error("❌ Please check the confirmation box first")
                    else:
                        with st.spinner("⏳ Adding schedule to Banner cart..."):
                            selected_term = st.session_state.get(
                                "selected_term", "202530"
                            )
                            success = apply_schedule_to_banner(
                                current_combo, selected_term
                            )

                            if success:
                                st.success("🎉 **Classes added to cart!**")
                                st.info(
                                    "🔍 Go to SAIT Banner to review and submit your registration"
                                )
                            else:
                                st.error("❌ **Failed to apply schedule**")
                                st.info(
                                    "💡 Please try registering manually through SAIT Banner or contact support"
                                )

    else:
        st.info("Add some classes first to generate schedules!")


# Drop Classes tab
def manual_schedule_editor():
    """Display and manage the user's existing schedule from Banner"""
    st.header("�️ Drop Classes")
    st.markdown("**View your current schedule and drop classes you no longer need.**")
    st.info(
        "ℹ️ Select the checkbox next to any class you want to drop, then click the 'Drop Selected Classes' button below."
    )
    st.warning(
        "⚠️ **Important:** You must have at least one class registered. The system won't allow you to drop all your classes."
    )

    # Check if authenticated
    if not st.session_state.get("banner_credentials"):
        st.warning("⚠️ You need to authenticate first to load your existing schedule")
        st.info("👆 Go to the sidebar and click 'Authenticate Now'")
        return

    selected_term = st.session_state.get("selected_term", "202530")

    # Add refresh button
    col1, col2 = st.columns([3, 1])
    with col1:
        st.markdown(f"**📅 Term:** {selected_term}")
    with col2:
        refresh = st.button("🔄 Refresh", use_container_width=True)

    # Fetch current registrations (calendar events)
    with st.spinner("Loading your current schedule from Banner..."):
        calendar_events = get_current_registrations(selected_term)

        # Also fetch full registration models and store in session state
        registration_models = get_registration_models_for_term(selected_term)
        if "registration_models" not in st.session_state:
            st.session_state.registration_models = {}
        st.session_state.registration_models = registration_models

    if not calendar_events:
        st.warning("📭 No registered classes found for this term")
        st.info(
            "If you have classes registered, try refreshing your authentication tokens"
        )
        return

    # Group events by CRN (each class has multiple events for different meeting times)
    # Also get section info from models if available
    classes_by_crn = {}
    for event in calendar_events:
        crn = event.get("crn", "")
        if not crn:
            continue

        if crn not in classes_by_crn:
            # Try to get section info from the model
            section = "A"  # Default
            if crn in registration_models:
                model = registration_models[crn]
                section = model.get("sequenceNumber", "A")

            classes_by_crn[crn] = {
                "title": event.get("title", "Unknown Course"),
                "subject": event.get("subject", ""),
                "courseNumber": event.get("courseNumber", ""),
                "section": section,
                "crn": crn,
                "term": event.get("term", selected_term),
                "events": [],
            }
        classes_by_crn[crn]["events"].append(event)

    st.markdown("---")
    st.markdown("### 📅 Your Current Schedule")

    if classes_by_crn:
        parsed_classes = []
        for crn, class_info in classes_by_crn.items():
            # Parse events into schedule format
            schedule_by_day = {}
            for event in class_info["events"]:
                start_str = event.get("start", "")
                end_str = event.get("end", "")

                if start_str and end_str:
                    try:
                        from datetime import datetime

                        # Fix timezone format: -0600 -> -06:00
                        def fix_timezone(dt_str):
                            if (
                                len(dt_str) >= 5
                                and dt_str[-5] in ["+", "-"]
                                and dt_str[-4:].isdigit()
                            ):
                                return dt_str[:-2] + ":" + dt_str[-2:]
                            return dt_str

                        start_str_fixed = fix_timezone(start_str)
                        end_str_fixed = fix_timezone(end_str)

                        # Parse ISO format datetime
                        start_dt = datetime.fromisoformat(start_str_fixed)
                        end_dt = datetime.fromisoformat(end_str_fixed)

                        day_name = start_dt.strftime("%A")
                        start_time = start_dt.strftime("%H:%M")
                        end_time = end_dt.strftime("%H:%M")

                        key = f"{day_name}_{start_time}_{end_time}"
                        if key not in schedule_by_day:
                            schedule_by_day[key] = {
                                "day": day_name,
                                "start_time": parse_time(start_time),
                                "end_time": parse_time(end_time),
                                "class_room": "TBA",
                            }
                    except Exception as e:
                        pass  # Skip events that can't be parsed

            if schedule_by_day:
                parsed_classes.append(
                    {
                        "name": f"{class_info['subject']} {class_info['courseNumber']}",
                        "group": class_info.get("section", "A"),
                        "schedule": list(schedule_by_day.values()),
                    }
                )

        if parsed_classes:
            time_slots = get_unique_time_slots(parsed_classes)
            timetable_html = display_timetable_html(
                parsed_classes, time_slots, parsed_classes
            )
            st.markdown(timetable_html, unsafe_allow_html=True)

            # Show class list below the timetable with drop checkboxes
            st.markdown("---")
            st.markdown("### 📚 Registered Classes")

            # Create table with checkboxes
            st.markdown(
                """
            <style>
                .class-table {
                    width: 100%;
                    margin: 20px 0;
                }
                .class-table th {
                    text-align: left;
                    padding: 10px;
                    background-color: #f0f2f6;
                }
                .class-table td {
                    padding: 10px;
                    border-bottom: 1px solid #e0e0e0;
                }
            </style>
            """,
                unsafe_allow_html=True,
            )

            # Store selected classes to drop
            if "classes_to_drop" not in st.session_state:
                st.session_state.classes_to_drop = []

            # Create table header
            col_check, col_course, col_title, col_section, col_crn = st.columns(
                [1, 2, 4, 1, 1.5]
            )
            with col_check:
                st.write("**Drop?**")
            with col_course:
                st.write("**Course**")
            with col_title:
                st.write("**Title**")
            with col_section:
                st.write("**Section**")
            with col_crn:
                st.write("**CRN**")

            # Create rows for each class
            for crn, class_info in classes_by_crn.items():
                subject = class_info["subject"]
                course_num = class_info["courseNumber"]
                title = class_info["title"]
                section = class_info.get("section", "A")

                col_check, col_course, col_title, col_section, col_crn_val = st.columns(
                    [1, 2, 4, 1, 1.5]
                )

                with col_check:
                    drop_selected = st.checkbox(
                        "Drop", key=f"drop_{crn}", label_visibility="collapsed"
                    )
                    if drop_selected and crn not in st.session_state.classes_to_drop:
                        st.session_state.classes_to_drop.append(crn)
                    elif not drop_selected and crn in st.session_state.classes_to_drop:
                        st.session_state.classes_to_drop.remove(crn)

                with col_course:
                    st.write(f"**{subject} {course_num}**")

                with col_title:
                    st.write(title)

                with col_section:
                    st.write(section)

                with col_crn_val:
                    st.write(crn)

            # Drop button
            if st.session_state.classes_to_drop:
                st.markdown("---")
                st.warning(
                    f"⚠️ {len(st.session_state.classes_to_drop)} class(es) selected for dropping"
                )

                col1, col2 = st.columns([1, 3])
                with col1:
                    if st.button("🗑️ Drop Selected Classes", type="primary"):
                        # Confirm drop
                        with st.spinner("Dropping classes..."):
                            result = drop_classes(
                                selected_term, st.session_state.classes_to_drop
                            )

                            if result.get("success"):
                                st.success(
                                    f"✅ Successfully dropped {len(st.session_state.classes_to_drop)} class(es)!"
                                )
                                st.balloons()
                                # Clear the selection
                                st.session_state.classes_to_drop = []
                                # Rerun to refresh the schedule
                                st.rerun()
                            else:
                                error_msg = result.get("error", "Unknown error")
                                st.error(f"❌ Failed to drop classes: {error_msg}")
                with col2:
                    st.caption(
                        "⚠️ Dropping a class releases your seat - another student may register for it."
                    )
                    st.caption(
                        "💡 You can re-register for the class if seats are still available."
                    )
        else:
            st.warning("Could not parse any class schedules")
    else:
        st.warning("No classes found")


# Calendar ICS Generator tab
def calendar_ics_generator():
    st.header("📅 Export Schedule to Calendar")
    st.markdown(
        "**Export your registered classes to a calendar file (.ics) that you can import into Google Calendar, Outlook, Apple Calendar, etc.**"
    )

    # Check if authenticated
    if not st.session_state.get("banner_credentials"):
        st.warning("⚠️ You need to authenticate first to load your schedule from Banner")
        st.info("👆 Go to the sidebar and click 'Authenticate Now'")
        st.markdown("---")
        st.markdown("### Or use manually added classes:")

        # Fallback to manual classes
        id = st.session_state.get("id", "default_user")
        classes = get_classes_from_db(id)
        if not classes:
            st.warning("No manually added classes found either.")
            return

        # Combine class names with their sections for display
        class_display_names = [
            f"{cls['name']} - Section {cls['group']}" for cls in classes
        ]

        selected_classes_display = st.multiselect(
            "Select classes to include in the final calendar",
            class_display_names,
            placeholder="Choose classes",
        )

        # Extract class names and groups from the selected display names
        selected_classes = [
            {
                "name": display.split(" - ")[0],
                "group": display.split(" - ")[1].replace("Section ", ""),
            }
            for display in selected_classes_display
        ]

        if selected_classes:
            start_date = st.date_input("Start date")
            end_date = st.date_input("End date of semester")

            if st.button("Generate calendar"):
                try:
                    ics_filename = generate_ics_file_for_classes(
                        selected_classes,
                        classes,
                        start_date.strftime("%Y-%m-%d"),
                        end_date.strftime("%Y-%m-%d"),
                    )

                    with open(ics_filename, "rb") as file:
                        st.download_button(
                            label="Download ICS file",
                            data=file,
                            file_name=ics_filename,
                            mime="text/calendar",
                        )
                except Exception as e:
                    st.error(f"An error occurred: {e}")
        else:
            st.warning("Please select at least one class to include in the calendar.")
        return

    # User is authenticated - load their registered schedule
    selected_term = st.session_state.get("selected_term", "202530")

    st.info(f"📚 Loading your registered classes for term {selected_term}...")

    # Fetch current registrations
    with st.spinner("Loading your schedule from Banner..."):
        calendar_events = get_current_registrations(selected_term)

        # Also fetch full registration models
        registration_models = get_registration_models_for_term(selected_term)
        if "registration_models" not in st.session_state:
            st.session_state.registration_models = {}
        st.session_state.registration_models = registration_models

    if not calendar_events:
        st.warning("📭 No registered classes found for this term")
        st.info(
            "If you have classes registered, try refreshing your authentication tokens"
        )
        return

    # Group events by CRN and convert to the format needed for ICS generation
    classes_by_crn = {}
    for event in calendar_events:
        crn = event.get("crn", "")
        if not crn:
            continue

        if crn not in classes_by_crn:
            # Try to get section info from the model
            section = "A"  # Default
            if crn in registration_models:
                model = registration_models[crn]
                section = model.get("sequenceNumber", "A")

            classes_by_crn[crn] = {
                "title": event.get("title", "Unknown Course"),
                "subject": event.get("subject", ""),
                "courseNumber": event.get("courseNumber", ""),
                "section": section,
                "crn": crn,
                "term": event.get("term", selected_term),
                "events": [],
            }
        classes_by_crn[crn]["events"].append(event)

    # Convert to the format expected by generate_ics_file_for_classes
    classes_for_ics = []
    for crn, class_info in classes_by_crn.items():
        schedule_by_day = {}

        # Get instructor and room from the registration model
        instructor = "TBA"
        if crn in registration_models:
            model = registration_models[crn]
            # Try to get instructor name
            faculty = model.get("faculty", [])
            if faculty and len(faculty) > 0:
                instructor_info = faculty[0]
                display_name = instructor_info.get("displayName", "")
                if display_name:
                    instructor = display_name

        for event in class_info["events"]:
            start_str = event.get("start", "")
            end_str = event.get("end", "")
            building = event.get("building", "")
            room = event.get("room", "")

            # Construct location from building and room
            location = "TBA"
            if building and room:
                location = f"{building} {room}"
            elif building:
                location = building
            elif room:
                location = room

            if start_str and end_str:
                try:
                    from datetime import datetime

                    # Fix timezone format: -0600 -> -06:00
                    def fix_timezone(dt_str):
                        if (
                            len(dt_str) >= 5
                            and dt_str[-5] in ["+", "-"]
                            and dt_str[-4:].isdigit()
                        ):
                            return dt_str[:-2] + ":" + dt_str[-2:]
                        return dt_str

                    start_str_fixed = fix_timezone(start_str)
                    end_str_fixed = fix_timezone(end_str)

                    # Parse ISO format datetime
                    start_dt = datetime.fromisoformat(start_str_fixed)
                    end_dt = datetime.fromisoformat(end_str_fixed)

                    day_name = start_dt.strftime("%A")
                    start_time = start_dt.strftime("%H:%M")
                    end_time = end_dt.strftime("%H:%M")

                    key = f"{day_name}_{start_time}_{end_time}"
                    if key not in schedule_by_day:
                        schedule_by_day[key] = {
                            "day": day_name,
                            "start_time": start_time,
                            "end_time": end_time,
                            "class_room": location,
                        }
                except Exception as e:
                    pass  # Skip events that can't be parsed

        if schedule_by_day:
            classes_for_ics.append(
                {
                    "name": f"{class_info['subject']} {class_info['courseNumber']}",
                    "group": class_info.get("section", "A"),
                    "schedule": list(schedule_by_day.values()),
                    "instructor": instructor,
                }
            )

    if not classes_for_ics:
        st.warning("Could not parse any class schedules")
        return

    st.success(f"✅ Found {len(classes_for_ics)} registered classes")

    # Show the classes
    st.markdown("### 📚 Your Registered Classes")
    for cls in classes_for_ics:
        st.write(f"• **{cls['name']}** - Section {cls['group']}")

    st.markdown("---")
    st.markdown("### 📅 Set Semester Dates")

    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("Start date", help="First day of classes")
    with col2:
        end_date = st.date_input("End date of semester", help="Last day of classes")

    if st.button("📥 Generate Calendar File", type="primary", use_container_width=True):
        try:
            # All classes are automatically selected since they're from Banner
            selected_classes = [
                {"name": cls["name"], "group": cls["group"]} for cls in classes_for_ics
            ]

            ics_filename = generate_ics_file_for_classes(
                selected_classes,
                classes_for_ics,
                start_date.strftime("%Y-%m-%d"),
                end_date.strftime("%Y-%m-%d"),
            )

            with open(ics_filename, "rb") as file:
                st.download_button(
                    label="⬇️ Download ICS file",
                    data=file,
                    file_name=ics_filename,
                    mime="text/calendar",
                    use_container_width=True,
                )

            st.success("✅ Calendar file generated successfully!")
            st.info(
                "💡 **Tip:** Import this .ics file into Google Calendar, Outlook, Apple Calendar, or any other calendar app to see your class schedule!"
            )

        except Exception as e:
            st.error(f"An error occurred: {e}")


if __name__ == "__main__":
    main()
